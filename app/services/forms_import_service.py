import re
import unicodedata
import zipfile
from pathlib import Path
from urllib.parse import unquote, urlparse

from fastapi import HTTPException

from app.config import (
    FORMS_IMPORT_MAX_FILES,
    FORMS_IMPORT_MAX_IMAGE_MB,
    FORMS_IMPORT_PREFIX,
    UPLOAD_DIR,
)
from app.database import create_person_with_embedding, get_person, list_units
from app.face_engine import InvalidImageError, MultipleFacesFoundError, NoFaceFoundError
from app.services.log_service import register_log
from app.services.people_service import process_person_photo_bytes


REQUIRED_COLUMNS = {
    "forms_id": "ID",
    "name": "Nome Completo",
    "unit": "Unidade",
    "role": "Cargo / Função",
    "photo": "Foto",
}
ACCEPTED_IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp"}


def filename_from_photo_url(photo_url: str) -> str:
    path = unquote(urlparse(str(photo_url)).path)
    return Path(path).name


def _normalize_text(value) -> str:
    normalized = unicodedata.normalize("NFKD", str(value or ""))
    ascii_text = "".join(char for char in normalized if not unicodedata.combining(char))
    return re.sub(r"\s+", " ", ascii_text).strip().lower()


def _slugify_filename(value: str) -> str:
    text = _normalize_text(value)
    text = re.sub(r"[^a-z0-9]+", "-", text)
    text = re.sub(r"-+", "-", text).strip("-")
    return text or "foto"


def _safe_cell(value):
    if value is None:
        return ""
    return str(value).strip()


def _row_error(row_number, forms_id=None, person_id=None, name=None, code="DATABASE_ERROR", message="", expected_filename=None):
    return {
        "row": row_number,
        "forms_id": forms_id,
        "person_id": person_id,
        "name": name,
        "code": code,
        "message": message,
        "expected_filename": expected_filename,
    }


def _empty_report():
    return {
        "total_rows": 0,
        "imported": 0,
        "photo_completed": 0,
        "ignored": 0,
        "errors_count": 0,
        "errors": [],
    }


def _load_rows(excel_path: Path):
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as error:
        raise HTTPException(status_code=500, detail="Dependência openpyxl não instalada.") from error

    try:
        workbook = load_workbook(excel_path, read_only=False, data_only=True)
    except Exception as error:
        raise HTTPException(status_code=400, detail="Planilha XLSX inválida.") from error

    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    header_row = next(rows, None)
    if not header_row:
        return None, []

    header_map = {}
    for index, header in enumerate(header_row):
        normalized = _normalize_text(header)
        if normalized and normalized not in header_map:
            header_map[normalized] = index

    missing = [
        label
        for label in REQUIRED_COLUMNS.values()
        if _normalize_text(label) not in header_map
    ]
    if missing:
        report = _empty_report()
        report["errors"].append(_row_error(
            1,
            code="MISSING_COLUMN",
            message=f"Colunas obrigatórias ausentes: {', '.join(missing)}.",
        ))
        report["errors_count"] = 1
        return None, report

    column_indexes = {
        key: header_map[_normalize_text(label)]
        for key, label in REQUIRED_COLUMNS.items()
    }

    data_rows = []
    for row_number, row in enumerate(rows, start=2):
        if not any(_safe_cell(value) for value in row):
            continue
        data_rows.append((row_number, {
            key: row[index] if index < len(row) else None
            for key, index in column_indexes.items()
        }))

    return column_indexes, data_rows


def _index_zip(zip_path: Path):
    try:
        archive = zipfile.ZipFile(zip_path)
    except zipfile.BadZipFile as error:
        raise HTTPException(status_code=400, detail="Arquivo ZIP inválido.") from error

    with archive:
        file_infos = [info for info in archive.infolist() if not info.is_dir()]
        if len(file_infos) > FORMS_IMPORT_MAX_FILES:
            raise HTTPException(status_code=400, detail="ZIP excede o limite de arquivos permitido.")

        index = {}
        ambiguous = set()
        for info in file_infos:
            filename = Path(info.filename).name
            if not filename:
                continue
            if filename.startswith(".") or filename in {"Thumbs.db", ".DS_Store"}:
                continue
            if info.filename.startswith("__MACOSX/"):
                continue

            extension = Path(filename).suffix.lower()
            if extension not in ACCEPTED_IMAGE_EXTENSIONS:
                continue

            key = filename.lower()
            if key in index:
                ambiguous.add(key)
            else:
                index[key] = info.filename

        return index, ambiguous


def _person_id_from_forms_id(raw_forms_id):
    value = _safe_cell(raw_forms_id)
    if not value:
        raise ValueError
    try:
        forms_id = int(float(value))
    except ValueError as error:
        raise ValueError from error
    return forms_id, f"{FORMS_IMPORT_PREFIX}-{forms_id:04d}"


def _unit_lookup():
    return {
        _normalize_text(unit["name"]): unit
        for unit in list_units()
    }


def _unique_image_path(person_id: str, name: str, extension: str) -> Path:
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    base = f"{person_id}-{_slugify_filename(name)}"
    candidate = UPLOAD_DIR / f"{base}{extension}"
    counter = 2
    while candidate.exists():
        candidate = UPLOAD_DIR / f"{base}-{counter}{extension}"
        counter += 1
    return candidate


def _read_zip_image(zip_path: Path, internal_path: str):
    with zipfile.ZipFile(zip_path) as archive:
        info = archive.getinfo(internal_path)
        max_bytes = FORMS_IMPORT_MAX_IMAGE_MB * 1024 * 1024
        if info.file_size > max_bytes:
            raise ValueError("PHOTO_TOO_LARGE")
        return archive.read(internal_path)


def _face_error_code(error: ValueError) -> str:
    if isinstance(error, InvalidImageError):
        return "INVALID_IMAGE"
    if isinstance(error, NoFaceFoundError):
        return "NO_FACE_FOUND"
    if isinstance(error, MultipleFacesFoundError):
        return "MULTIPLE_FACES_FOUND"
    return "INVALID_IMAGE"


def _error_message(code: str) -> str:
    messages = {
        "INVALID_FORMS_ID": "ID do Forms inválido.",
        "MISSING_NAME": "Nome completo não informado.",
        "MISSING_UNIT": "Unidade não informada.",
        "UNIT_NOT_FOUND": "Unidade não encontrada.",
        "MISSING_PHOTO_URL": "URL da foto não informada.",
        "PHOTO_NOT_FOUND": "Foto não encontrada no ZIP.",
        "PHOTO_NAME_AMBIGUOUS": "Nome de foto duplicado dentro do ZIP.",
        "INVALID_PHOTO_EXTENSION": "Extensão de foto inválida.",
        "PHOTO_TOO_LARGE": "Foto excede o tamanho máximo permitido.",
        "INVALID_IMAGE": "Imagem inválida.",
        "NO_FACE_FOUND": "Nenhum rosto encontrado na foto.",
        "MULTIPLE_FACES_FOUND": "Mais de um rosto encontrado na foto.",
        "PERSON_DATA_CONFLICT": "Pessoa já existe com nome diferente.",
        "DATABASE_ERROR": "Erro ao gravar no banco de dados.",
        "FILE_WRITE_ERROR": "Erro ao gravar a imagem.",
    }
    return messages.get(code, "Erro na importação.")


def import_forms_people(excel_path: Path, zip_path: Path, face_engine) -> dict:
    _, loaded = _load_rows(excel_path)
    if isinstance(loaded, dict):
        return loaded

    zip_index, ambiguous_names = _index_zip(zip_path)
    units = _unit_lookup()
    report = _empty_report()
    report["total_rows"] = len(loaded)

    for row_number, row in loaded:
        raw_forms_id = _safe_cell(row["forms_id"])
        name = _safe_cell(row["name"])
        unit_name = _safe_cell(row["unit"])
        role = _safe_cell(row["role"]) or None
        photo_url = _safe_cell(row["photo"])
        forms_id = None
        person_id = None
        expected_filename = None

        def add_error(code):
            error = _row_error(
                row_number,
                forms_id=forms_id or raw_forms_id or None,
                person_id=person_id,
                name=name or None,
                code=code,
                message=_error_message(code),
                expected_filename=expected_filename,
            )
            report["errors"].append(error)
            register_log("IMPORT_FORMS_ERROR", person_id, f"Linha {row_number}: {code}")

        try:
            forms_id, person_id = _person_id_from_forms_id(raw_forms_id)
        except ValueError:
            add_error("INVALID_FORMS_ID")
            continue

        if not name:
            add_error("MISSING_NAME")
            continue
        if not unit_name:
            add_error("MISSING_UNIT")
            continue
        unit = units.get(_normalize_text(unit_name))
        if not unit:
            add_error("UNIT_NOT_FOUND")
            continue
        if not photo_url:
            add_error("MISSING_PHOTO_URL")
            continue

        expected_filename = filename_from_photo_url(photo_url)
        extension = Path(expected_filename).suffix.lower()
        if extension not in ACCEPTED_IMAGE_EXTENSIONS:
            add_error("INVALID_PHOTO_EXTENSION")
            continue

        zip_key = expected_filename.lower()
        if zip_key in ambiguous_names:
            add_error("PHOTO_NAME_AMBIGUOUS")
            continue
        internal_path = zip_index.get(zip_key)
        if not internal_path:
            add_error("PHOTO_NOT_FOUND")
            continue

        existing_person = get_person(person_id)
        if existing_person and _normalize_text(existing_person["name"]) != _normalize_text(name):
            add_error("PERSON_DATA_CONFLICT")
            continue
        if existing_person and int(existing_person.get("photos_count") or 0) > 0:
            report["ignored"] += 1
            continue

        try:
            image_bytes = _read_zip_image(zip_path, internal_path)
        except ValueError as error:
            add_error(str(error))
            continue

        if existing_person:
            try:
                process_person_photo_bytes(person_id, image_bytes, expected_filename, face_engine)
            except HTTPException as error:
                code = "INVALID_IMAGE"
                detail = str(error.detail)
                if "Nenhum rosto" in detail:
                    code = "NO_FACE_FOUND"
                elif "Mais de um rosto" in detail:
                    code = "MULTIPLE_FACES_FOUND"
                add_error(code)
                continue

            report["photo_completed"] += 1
            register_log("IMPORT_FORMS_PHOTO_COMPLETED", person_id, f"Foto importada do Forms: {name}")
            continue

        try:
            embedding = face_engine.get_embedding(image_bytes)
        except ValueError as error:
            add_error(_face_error_code(error))
            continue

        image_path = _unique_image_path(person_id, name, extension)
        try:
            image_path.write_bytes(image_bytes)
        except OSError:
            add_error("FILE_WRITE_ERROR")
            continue

        try:
            create_person_with_embedding(
                person_id=person_id,
                name=name,
                embedding=embedding,
                image_path=str(image_path),
                unit_id=unit["id"],
                unit=unit["name"],
                role=role,
            )
        except Exception:
            image_path.unlink(missing_ok=True)
            add_error("DATABASE_ERROR")
            continue

        report["imported"] += 1
        register_log("IMPORT_FORMS_PERSON", person_id, f"Pessoa importada do Forms com consentimento confirmado: {name}")

    report["errors_count"] = len(report["errors"])
    return report
